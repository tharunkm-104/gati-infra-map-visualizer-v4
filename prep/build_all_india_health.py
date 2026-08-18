#!/usr/bin/env python3
"""
Build all-India health infrastructure (NABH + NMC + INC) from:
  - nabh_raw_rows.jsonl (raw NABH accreditation rows, already scraped)
  - City_German_Language_and_Medical_Infra_Mapping.xlsx sheets:
      'NMC- Medical Colleges', 'INC - Nursing Colleges'

Outputs:
  data/all-india-health-points.json  -- one row per facility, ownership tagged
  data/all-india-states.json         -- per-state totals (govt/private split)

NABH facilities have no per-facility ownership field in the source data
(Data Sources Selected marks the category "Both" but does not distinguish
individual rows), so NABH points carry ownership: "Not specified".
"""
import html as ihtml, json, re, sys
from collections import defaultdict
import openpyxl

XLSX = "/mnt/user-data/uploads/City_German_Language_and_Medical_Infra_Mapping.xlsx"
NABH_JSONL = "/mnt/user-data/uploads/nabh_raw_rows.jsonl"
OUT_POINTS = "data/all-india-health-points.json"
OUT_STATES = "data/all-india-states.json"

def _clean(s):
    return ihtml.unescape(re.sub(r"\s+", " ", s or "")).strip()

STATES = {
 'andhra pradesh':'Andhra Pradesh','arunachal pradesh':'Arunachal Pradesh','assam':'Assam',
 'bihar':'Bihar','chhattisgarh':'Chhattisgarh','chattisgarh':'Chhattisgarh','goa':'Goa',
 'gujarat':'Gujarat','haryana':'Haryana','himachal pradesh':'Himachal Pradesh',
 'jharkhand':'Jharkhand','karnataka':'Karnataka','kerala':'Kerala','madhya pradesh':'Madhya Pradesh',
 'maharashtra':'Maharashtra','manipur':'Manipur','meghalaya':'Meghalaya','mizoram':'Mizoram',
 'nagaland':'Nagaland','odisha':'Odisha','orissa':'Odisha','punjab':'Punjab','rajasthan':'Rajasthan',
 'sikkim':'Sikkim','tamil nadu':'Tamil Nadu','tamilnadu':'Tamil Nadu','telangana':'Telangana',
 'tripura':'Tripura','uttar pradesh':'Uttar Pradesh','uttarakhand':'Uttarakhand',
 'uttaranchal':'Uttarakhand','west bengal':'West Bengal',
 'andaman and nicobar islands':'Andaman and Nicobar Islands','andaman and nicobar':'Andaman and Nicobar Islands',
 'andaman & nicobar':'Andaman and Nicobar Islands','chandigarh':'Chandigarh',
 'dadra and nagar haveli and daman and diu':'Dadra and Nagar Haveli and Daman and Diu',
 'dadra and nagar haveli':'Dadra and Nagar Haveli and Daman and Diu',
 'daman and diu':'Dadra and Nagar Haveli and Daman and Diu','daman & diu':'Dadra and Nagar Haveli and Daman and Diu',
 'new delhi':'Delhi','delhi':'Delhi','jammu and kashmir':'Jammu and Kashmir',
 'jammu & kashmir':'Jammu and Kashmir','j&k':'Jammu and Kashmir','ladakh':'Ladakh',
 'lakshadweep':'Lakshadweep','puducherry':'Puducherry','pondicherry':'Puducherry',
}
_ALIASES = sorted(STATES.keys(), key=len, reverse=True)

CITY_STATE = {
 'bangalore':'Karnataka','bengaluru':'Karnataka','mysore':'Karnataka','mysuru':'Karnataka',
 'mangalore':'Karnataka','hubli':'Karnataka','belgaum':'Karnataka',
 'mumbai':'Maharashtra','pune':'Maharashtra','nagpur':'Maharashtra','nashik':'Maharashtra',
 'thane':'Maharashtra','aurangabad':'Maharashtra','navi mumbai':'Maharashtra','kolhapur':'Maharashtra',
 'chennai':'Tamil Nadu','coimbatore':'Tamil Nadu','madurai':'Tamil Nadu','trichy':'Tamil Nadu',
 'tiruchirappalli':'Tamil Nadu','salem':'Tamil Nadu','erode':'Tamil Nadu','vellore':'Tamil Nadu',
 'hyderabad':'Telangana','secunderabad':'Telangana','warangal':'Telangana',
 'kolkata':'West Bengal','calcutta':'West Bengal','howrah':'West Bengal','siliguri':'West Bengal',
 'ahmedabad':'Gujarat','surat':'Gujarat','vadodara':'Gujarat','rajkot':'Gujarat','baroda':'Gujarat',
 'jaipur':'Rajasthan','jodhpur':'Rajasthan','udaipur':'Rajasthan','kota':'Rajasthan','ajmer':'Rajasthan',
 'lucknow':'Uttar Pradesh','kanpur':'Uttar Pradesh','noida':'Uttar Pradesh','ghaziabad':'Uttar Pradesh',
 'varanasi':'Uttar Pradesh','agra':'Uttar Pradesh','meerut':'Uttar Pradesh','allahabad':'Uttar Pradesh',
 'prayagraj':'Uttar Pradesh','gorakhpur':'Uttar Pradesh','bareilly':'Uttar Pradesh',
 'patna':'Bihar','gaya':'Bihar','bhagalpur':'Bihar','muzaffarpur':'Bihar',
 'gurgaon':'Haryana','gurugram':'Haryana','faridabad':'Haryana','panipat':'Haryana','karnal':'Haryana',
 'rohtak':'Haryana','hisar':'Haryana','ambala':'Haryana',
 'indore':'Madhya Pradesh','bhopal':'Madhya Pradesh','gwalior':'Madhya Pradesh','jabalpur':'Madhya Pradesh',
 'ujjain':'Madhya Pradesh','kochi':'Kerala','cochin':'Kerala','ernakulam':'Kerala',
 'thiruvananthapuram':'Kerala','trivandrum':'Kerala','kozhikode':'Kerala','calicut':'Kerala',
 'thrissur':'Kerala','kollam':'Kerala','kottayam':'Kerala',
 'bhubaneswar':'Odisha','cuttack':'Odisha','rourkela':'Odisha',
 'raipur':'Chhattisgarh','bhilai':'Chhattisgarh','bilaspur':'Chhattisgarh',
 'ranchi':'Jharkhand','jamshedpur':'Jharkhand','dhanbad':'Jharkhand',
 'guwahati':'Assam','dibrugarh':'Assam','mohali':'Punjab','panchkula':'Haryana',
 'ludhiana':'Punjab','amritsar':'Punjab','jalandhar':'Punjab','patiala':'Punjab',
 'dehradun':'Uttarakhand','haridwar':'Uttarakhand','rishikesh':'Uttarakhand',
 'visakhapatnam':'Andhra Pradesh','vizag':'Andhra Pradesh','vijayawada':'Andhra Pradesh',
 'guntur':'Andhra Pradesh','tirupati':'Andhra Pradesh','nellore':'Andhra Pradesh',
 'shimla':'Himachal Pradesh','srinagar':'Jammu and Kashmir','jammu':'Jammu and Kashmir',
 'panaji':'Goa','panjim':'Goa','margao':'Goa','imphal':'Manipur','shillong':'Meghalaya',
 'agartala':'Tripura','aizawl':'Mizoram','kohima':'Nagaland','itanagar':'Arunachal Pradesh',
 'gangtok':'Sikkim',
}

def extract_geo(addr):
    a = _clean(addr); low = a.lower()
    pin = re.search(r'\b(\d{6})\b', a)
    pin = pin.group(1) if pin else ""
    best = None
    for al in _ALIASES:
        for m in re.finditer(r'(?<![a-z])' + re.escape(al) + r'(?![a-z])', low):
            if best is None or m.start() > best[0]:
                best = (m.start(), al)
    state = STATES[best[1]] if best else ""
    city = ""
    if not state:
        for cty, st in CITY_STATE.items():
            if re.search(r'(?<![a-z])' + re.escape(cty) + r'(?![a-z])', low):
                state, city = st, cty.title()
                return state, city, pin
    if state:
        parts = [p.strip() for p in a.split(",") if p.strip()]
        si = next((i for i, p in enumerate(parts) if best[1] in p.lower()), None)
        if si is not None and si - 1 >= 0:
            cand = re.sub(r'\b\d{3,6}\b', '', parts[si - 1])
            cand = re.sub(r'[-\u2013]\s*\d+.*$', '', cand).strip(" -\u2013")
            cand = re.sub(r'\s+', ' ', cand)
            if cand.isdigit() and si - 2 >= 0:
                cand = re.sub(r'\b\d{3,6}\b', '', parts[si - 2]).strip(" -\u2013")
            city = cand.title() if cand and not cand.isdigit() else ""
    return state, city, pin

ELIGIBLE_TYPE = {
    'Hospitals': True, 'Entry-Level Hospitals': True, 'SHCO': True, 'Entry-Level SHCO': True,
    'Nursing Home / Nursing Excellence': True, 'Care Home': True, 'Eye Care Organisation': True,
    'Allopathic Clinics': True, 'Dental Facilities': True,
    'CGHS/ECHS Empaneled Hospital': True, 'Eye Care Empanelment (CGHS/ECHS)': True,
    'Dental Empanelment (CGHS/ECHS)': True,
    'Emergency Department': True, 'Stroke Centre': True,
    'AYUSH Hospitals': False, 'Entry Level Ayush Hospital': False, 'Entry Level Ayush Centre': False,
    'AYUSH Empanelment (CGHS/ECHS)': False, 'Panchakarma Clinics': False, 'Blood Bank': False,
    'Medical Imaging Services': False, 'Medical Laboratory': False, 'Dialysis Centre': False,
    'Diagnostic/Lab Empanelment (CGHS/ECHS)': False, 'Entry-Level Dental Clinics': False,
    'Ethics Committee (non-clinical)': False, 'Medical Value Travel (MVTF)': False,
    'Digital Health': False, 'Climate/Sustainability': False,
    'Unspecified': False,
}
FOLDER_MAP = {
    'hospitals': 'Hospitals', 'hco': 'Hospitals',
    'entry-level hospitals': 'Entry-Level Hospitals', 'entry level hospitals': 'Entry-Level Hospitals',
    'shco': 'SHCO', 'entry-level shco': 'Entry-Level SHCO', 'entry level shco': 'Entry-Level SHCO',
    'entry-level scho': 'Entry-Level SHCO', 'entry level scho': 'Entry-Level SHCO',
    'nursing excellence': 'Nursing Home / Nursing Excellence', 'nursing': 'Nursing Home / Nursing Excellence',
    'care home': 'Care Home', 'eye care': 'Eye Care Organisation', 'eye care organisation': 'Eye Care Organisation',
    'allopathic clinics': 'Allopathic Clinics', 'allopathic clinics 2nd edition': 'Allopathic Clinics',
    'dental facilities': 'Dental Facilities', 'dental healthcare service providers': 'Dental Facilities',
    'entry level dental clinics': 'Entry-Level Dental Clinics', 'entry-level dental clinics': 'Entry-Level Dental Clinics',
    'ayush hospitals': 'AYUSH Hospitals', 'entry level ayush hospitals': 'Entry Level Ayush Hospital',
    'entry level ayush centre': 'Entry Level Ayush Centre', 'entry level ayush center': 'Entry Level Ayush Centre',
    'panchakarma': 'Panchakarma Clinics', 'panchakarma clinics': 'Panchakarma Clinics',
    'blood bank': 'Blood Bank', 'blood banks': 'Blood Bank', 'blood centre': 'Blood Bank',
    'mis': 'Medical Imaging Services', 'medical imaging services': 'Medical Imaging Services',
    'medical laboratory': 'Medical Laboratory', 'laboratory': 'Medical Laboratory',
    'dialysis': 'Dialysis Centre', 'dialysis centre': 'Dialysis Centre',
    'clinical trial': 'Ethics Committee (non-clinical)', 'ethics committee': 'Ethics Committee (non-clinical)',
    'emergency': 'Emergency Department', 'stroke': 'Stroke Centre',
    'digital health': 'Digital Health', 'ccrs': 'Climate/Sustainability',
    'mvtf': 'Medical Value Travel (MVTF)', 'medical value travel': 'Medical Value Travel (MVTF)',
}
PREFIX_MAP = {
    'PESHCO': 'Entry-Level SHCO', 'PEHCO': 'Entry-Level Hospitals', 'PEH': 'Entry-Level Hospitals',
    'ESHCO': 'Entry-Level SHCO', 'ESCHO': 'Entry-Level SHCO', 'EH': 'Entry-Level Hospitals',
    'SHCO': 'SHCO', 'SCHO': 'SHCO', 'HCO': 'Hospitals', 'HOS': 'CGHS/ECHS Empaneled Hospital',
    'ECO': 'Eye Care Organisation', 'EYE': 'Eye Care Organisation',
    'BB': 'Blood Bank', 'BC': 'Blood Bank', 'AC': 'Allopathic Clinics',
    'MIS': 'Medical Imaging Services', 'AH': 'AYUSH Hospitals', 'PC': 'Panchakarma Clinics',
    'CD': 'Dialysis Centre', 'DHSP': 'Dental Facilities', 'DHC': 'Digital Health',
    'DHS': 'Digital Health', 'DH': 'Digital Health', 'DEN': 'Dental Facilities', 'DE': 'Dental Facilities',
    'ELDC': 'Entry-Level Dental Clinics', 'EDC': 'Entry-Level Dental Clinics', 'D': 'Dental Facilities',
    'EAC': 'Entry Level Ayush Centre', 'EAH': 'Entry Level Ayush Hospital',
    'ML': 'Medical Laboratory', 'MLP': 'Medical Laboratory', 'LAB': 'Medical Laboratory',
    'ED': 'Emergency Department', 'ST': 'Stroke Centre', 'CCRS': 'Climate/Sustainability',
    'EC-CT': 'Ethics Committee (non-clinical)', 'EC': 'Ethics Committee (non-clinical)',
    'CT': 'Ethics Committee (non-clinical)', 'MT': 'Medical Value Travel (MVTF)',
    'CH': 'Care Home', 'N': 'Nursing Home / Nursing Excellence', 'H': 'Hospitals',
}
_PREFIX_KEYS = sorted(PREFIX_MAP.keys(), key=len, reverse=True)

def _underlying_from_name(name):
    n = " " + name.lower() + " "
    if re.search(r'ayurved|homoeo|homeo|unani|siddha|\byoga\b|naturopath|panchakarma|\bayush\b', n):
        return 'AYUSH Hospitals'
    if re.search(r'dental|dentist|orthodon', n):
        return 'Dental Empanelment (CGHS/ECHS)'
    if re.search(r'\beye\b|ophthalm|netralay|nethralay|vision|\bdrishti\b', n):
        return 'Eye Care Empanelment (CGHS/ECHS)'
    if re.search(r'diagnost|laborator|\blab\b|patholog|\bscan\b|imaging|radiolog|\bmri\b|\bct\b', n):
        return 'Diagnostic/Lab Empanelment (CGHS/ECHS)'
    if re.search(r'dialysis|nephro', n):
        return 'Dialysis Centre'
    if re.search(r'blood\s*bank|blood\s*cent', n):
        return 'Blood Bank'
    return 'CGHS/ECHS Empaneled Hospital'

def classify(cert, acc, nomen, name):
    m = re.search(r'/AccreditedList/([^/]+)/', cert or "")
    if m:
        key = re.sub(r'\s+', ' ', ihtml.unescape(m.group(1))).strip().lower()
        key = re.sub(r'[_%20]+', ' ', key).strip()
        if key in FOLDER_MAP:
            return FOLDER_MAP[key]
        for fk, fv in FOLDER_MAP.items():
            if fk in key:
                return fv
    accU = (acc or "").strip().upper()
    if re.match(r'^\d{4}[-/]\d', accU) or accU.startswith('HOS/'):
        return _underlying_from_name(name)
    nomU = (nomen or "").strip().upper()
    for k in _PREFIX_KEYS:
        if nomU == k or nomU.startswith(k):
            return PREFIX_MAP[k]
    token = re.split(r'[-/ ]', accU, 1)[0] if accU else ""
    for k in _PREFIX_KEYS:
        if accU.startswith(k + '-') or accU.startswith(k + '/') or token == k:
            return PREFIX_MAP[k]
    return 'Unspecified'

def name_key(s):
    s = ihtml.unescape(s or "").lower()
    s = re.sub(r'[^a-z0-9 ]', ' ', s)
    s = re.sub(r'\b(pvt|private|ltd|limited|the|a|unit|of|and|dr|shri|sri|m\/s)\b', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def facility_key(nm_key, state, pin, city):
    loc = pin if pin else (city.lower() if city else state.lower())
    return (nm_key, state.lower(), loc)

def build_nabh_points():
    seen_acc, raw = set(), []
    with open(NABH_JSONL, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                r = json.loads(line)
            except Exception:
                continue
            a = (r.get("acc") or "").strip().upper()
            k = a or ("NOACC::" + name_key(r.get("name", "")) + "::" + (r.get("addr", "")[:40]))
            if k in seen_acc:
                continue
            seen_acc.add(k)
            raw.append(r)

    for r in raw:
        r["state"], r["city"], r["pin"] = extract_geo(r.get("addr", ""))
        r["ptype"] = classify(r.get("cert", ""), r.get("acc", ""), r.get("nomen", ""), r.get("name", ""))

    fac = {}
    for r in raw:
        nk = name_key(r["name"]) or (r.get("acc", "").upper())
        key = facility_key(nk, r["state"], r["pin"], r["city"])
        f = fac.get(key)
        if f is None:
            f = fac[key] = {"rep": r, "types": []}
        if len(r.get("addr", "")) > len(f["rep"].get("addr", "")):
            f["rep"] = r
        if r["ptype"] not in f["types"]:
            f["types"].append(r["ptype"])

    points = []
    dropped_geo = 0
    for f in fac.values():
        eligible = any(ELIGIBLE_TYPE.get(t, False) for t in f["types"])
        if not eligible:
            continue
        rep = f["rep"]
        try:
            lat, lng = float(rep.get("lat")), float(rep.get("lng"))
        except (TypeError, ValueError):
            dropped_geo += 1
            continue
        if not rep["state"]:
            dropped_geo += 1
            continue
        points.append({
            "name": rep["name"], "subtype": "NABH Accredited Health Facility",
            "programTypes": f["types"], "ownership": "Not specified",
            "state": rep["state"], "city": rep["city"] or "", "address": rep.get("addr", ""),
            "latitude": lat, "longitude": lng, "source": "NABH", "coordinateStatus": "source",
        })
    sys.stderr.write(f"NABH: {len(raw)} raw rows -> {len(fac)} unique facilities -> "
                     f"{len(points)} eligible+geocoded ({dropped_geo} eligible-but-ungeocoded dropped)\n")
    return points

def _safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def norm_ownership(raw):
    raw = (raw or "").strip()
    low = raw.lower()
    if "government" in low:
        return "Government"
    if "private" in low or "trust" in low or "society" in low:
        return "Private"
    return "Not specified"

def build_nmc_points(wb):
    ws = wb["NMC- Medical Colleges"]
    points = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name:
            continue
        state, city = row[2], row[3]
        mgmt = row[6]
        lat, lng = _safe_float(row[12]), _safe_float(row[13])
        if lat is None or lng is None:
            skipped += 1
            continue
        points.append({
            "name": str(name), "subtype": "NMC Medical College",
            "ownership": norm_ownership(mgmt),
            "state": str(state or ""), "city": str(city or ""), "address": str(name),
            "latitude": lat, "longitude": lng, "source": "NMC", "coordinateStatus": "source",
        })
    sys.stderr.write(f"NMC: {len(points)} geocoded ({skipped} skipped, bad coordinates)\n")
    return points

def build_inc_points(wb):
    ws = wb["INC - Nursing Colleges"]
    points = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1]
        if not name:
            continue
        state, city = row[2], row[3]
        sector = row[6]
        lat, lng = _safe_float(row[8]), _safe_float(row[9])
        if lat is None or lng is None:
            skipped += 1
            continue
        points.append({
            "name": str(name), "subtype": "INC Nursing College",
            "ownership": norm_ownership(sector),
            "state": str(state or ""), "city": str(city or ""), "address": str(name),
            "latitude": lat, "longitude": lng, "source": "INC", "coordinateStatus": "source",
        })
    sys.stderr.write(f"INC: {len(points)} geocoded ({skipped} skipped, bad coordinates)\n")
    return points

STATE_NORM = {
    "andaman & nicobar": "Andaman and Nicobar Islands", "andaman & nicobar islands": "Andaman and Nicobar Islands",
    "jammu & kashmir": "Jammu and Kashmir", "dadra & nagar haveli": "Dadra and Nagar Haveli and Daman and Diu",
    "daman & diu": "Dadra and Nagar Haveli and Daman and Diu", "orissa": "Odisha",
    "pondicherry": "Puducherry", "nct of delhi": "Delhi", "delhi (nct)": "Delhi",
}

def norm_state(s):
    s = (s or "").strip()
    return STATE_NORM.get(s.lower(), s)

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    nabh = build_nabh_points()
    nmc = build_nmc_points(wb)
    inc = build_inc_points(wb)

    points = nabh + nmc + inc
    for p in points:
        p["state"] = norm_state(p["state"])

    with open(OUT_POINTS, "w", encoding="utf-8") as f:
        json.dump(points, f, ensure_ascii=False)

    by_state = defaultdict(lambda: {"nabh_facilities": 0, "nmc_colleges": 0, "inc_colleges": 0,
                                     "government": 0, "private": 0, "not_specified": 0,
                                     "lat_sum": 0.0, "lng_sum": 0.0, "n": 0})
    for p in points:
        st = p["state"] or "(unspecified)"
        b = by_state[st]
        if p["subtype"] == "NABH Accredited Health Facility":
            b["nabh_facilities"] += 1
        elif p["subtype"] == "NMC Medical College":
            b["nmc_colleges"] += 1
        elif p["subtype"] == "INC Nursing College":
            b["inc_colleges"] += 1
        if p["ownership"] == "Government":
            b["government"] += 1
        elif p["ownership"] == "Private":
            b["private"] += 1
        else:
            b["not_specified"] += 1
        b["lat_sum"] += p["latitude"]; b["lng_sum"] += p["longitude"]; b["n"] += 1

    states_out = []
    for st, b in sorted(by_state.items()):
        total = b["nabh_facilities"] + b["nmc_colleges"] + b["inc_colleges"]
        states_out.append({
            "state": st, "nabh_facilities": b["nabh_facilities"], "nmc_colleges": b["nmc_colleges"],
            "inc_colleges": b["inc_colleges"], "total": total,
            "government": b["government"], "private": b["private"], "not_specified": b["not_specified"],
            "latitude": b["lat_sum"] / b["n"], "longitude": b["lng_sum"] / b["n"],
        })

    with open(OUT_STATES, "w", encoding="utf-8") as f:
        json.dump(states_out, f, ensure_ascii=False, indent=1)

    sys.stderr.write(f"\nTOTAL all-India health points: {len(points)} "
                     f"(NABH {len(nabh)}, NMC {len(nmc)}, INC {len(inc)})\n")
    sys.stderr.write(f"states/UTs: {len(states_out)}\n")

if __name__ == "__main__":
    main()
